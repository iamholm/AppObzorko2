import re
import openpyxl
import os

class ImprovedAddressProcessor:
    def __init__(self):
        # Словарь, сопоставляющий названия улиц с их типами
        # Содержит только улицы Санкт-Петербурга, которые нужно обрабатывать
        self.street_types = {
            "Академика Байкова": "ул.",
            "Академика Глушко": "аллея",
            "Академика Константинова": "ул.",
            "Академика Лебедева": "ул.",
            "Амурская": "ул.",
            "Антоновская": "ул.",
            "Арсенальная": "наб.",
            "Арсенальная": "ул.",
            "Бестужевская": "ул.",
            "Бобруйская": "ул.",
            "Богословская": "ул.",
            "Боткинская": "ул.",
            "Брюсовская": "ул.",
            "Брянцева": "ул.",
            "Бутлерова": "ул.",
            "Вавиловых": "ул.",
            "Васенко": "ул.",
            "Ватутина": "ул.",
            "Веденеева": "ул.",
            "Верности": "ул.",
            "Верхняя": "ул.",
            "Герасимовская": "ул.",
            "Гжатская": "ул.",
            "Гидротехников": "ул.",
            "Гражданский": "пр.",
            "Демьяна Бедного": "ул.",
            "Жукова": "ул.",
            "Замшина": "ул.",
            "Карпинского": "ул.",
            "Киришская": "ул.",
            "Ключевая": "ул.",
            "Комиссара Смирнова": "ул.",
            "Комсомола": "ул.",
            "Кондратьевский": "пр.",
            "Культуры": "пр.",
            "Кушелевская": "дор.",
            "Лабораторная": "ул.",
            "Лабораторный": "пр.",
            "Лесной": "пр.",
            "Литовская": "ул.",
            "Лужская": "ул.",
            "Луначарского": "пр.",
            "Маршала Блюхера": "пр.",
            "Менделеевская": "ул.",
            "Меншиковский": "пр.",
            "Металлистов": "пр.",
            "Мечникова": "пр.",
            "Минеральная": "ул.",
            "Михайлова": "ул.",
            "Нартовская": "ул.",
            "Науки": "пр.",
            "Нейшлотский": "пер.",
            "Непокорённых": "пр.",
            "Новороссийская": "ул.",
            "Обручевых": "ул.",
            "Ольги Форш": "ул.",
            "Печорская": "ул.",
            "Пискарёвский": "пр.",
            "Политехническая": "ул.",
            "Полюстровский": "пр.",
            "Просвещения": "пр.",
            "Руставели": "ул.",
            "Свердловская": "наб.",
            "Светлановский": "пр.",
            "Северный": "пр.",
            "Сибирская": "ул.",
            "Софьи Ковалевской": "ул.",
            "Старо-Муринская": "ул.",
            "Старцева": "ул.",
            "Суздальский": "пр.",
            "Тимуровская": "ул.",
            "Тихорецкий": "пр.",
            "Токсовская": "ул.",
            "Архитектора Баранова": "ул.",
            "Рериха": "ул.",
            "Усыскина": "пер.",
            "Учительская": "ул.",
            "Ушинского": "ул.",
            "Фаворского": "ул.",
            "Федосеенко": "ул.",
            "Феодосийская": "ул.",
            "Финский": "пер.",
            "Хлопина": "ул.",
            "Черкасова": "ул.",
            "Чичуринский": "пер.",
            "Чугунная": "ул."
        }
        
        # Создаем список названий улиц из ключей словаря
        self.street_names = list(self.street_types.keys())
        
        # Префиксы улиц для удаления
        self.street_prefixes = [
            "пр\. ", "пр\.", "пр ", "пр",
            "ул\. ", "ул\.", "ул ", "ул",
            "проспект ", "проспект",
            "улица ", "улица",
            "аллея ", "б-р ", "бульвар ", "дор\. ", "дорога ",
            "наб\. ", "набережная ", "пер\. ", "переулок ",
            "пл\. ", "площадь ", "проезд ", "просп\. ", "просп"
        ]
        
        # Город
        self.city_prefixes = [
            "г\.СПб", "г\. СПб", "СПб", "Санкт-Петербург", 
            "г\.Санкт-Петербург", "г\. Санкт-Петербург"
        ]
        
        # Паттерны для поиска информации о доме - от самого специфического к самому общему
        self.house_patterns = [
            # НОВЫЕ ДОБАВЛЯЕМЫЕ ПАТТЕРНЫ
            # 1. "54-а-121" => "54А-121" (дом-литера-квартира через дефисы с маленькой буквой)
            r'\s*(\d+)-([а-я])-(\d+)',
            
            # 2. "д. 8/3/А кв. 189" => "8-3А-189" (дом/корпус/литера квартира)
            r'\s*д\.?\s*(\d+)\/(\d+)\/([А-Яа-я])\s+кв\.?\s*(\d+)',
            
            # 3. "110- А-422" => "110А-422" (дом-пробел-литера-квартира с пробелами)
            r'\s*(\d+)-\s*([А-Яа-я])-(\d+)',
            
            # 4. "16 лит. А кв. 43" => "16А-43" (дом литера квартира)
            r'\s*(\d+)\s+лит\.?\s+([А-Яа-я])\s+кв\.?\s*(\d+)',
            
            # 1. Самые специфичные форматы с наибольшим количеством компонентов
            # "д. 14, корп. 1, лит. А, кв. 93" - дом, корпус, литера, квартира с запятыми
            r'\s*д\.?\s*(\d+),\s*корп\.?\s*(\d+),\s*лит\.?\s*([А-Яа-я]),\s*кв\.?\s*(\d+)',
            
            # "27-2-А-17" - дом-корпус-литера-квартира через дефисы
            r'\s*(\d+)-(\d+)-([А-Яа-я])-(\d+)',
            
            # "5-2А-2" - дом-корпуслитера-квартира через дефисы
            r'\s*(\d+)-(\d+)([А-Яа-я])-(\d+)',
            
            # "34 к.1, лит. А, кв. 64" - дом, корпус, литера, квартира с запятыми
            r'\s*(\d+)\s*к\.?\s*(\d+),\s*лит\.?\s*([А-Яа-я]),\s*кв\.?\s*(\d+)',
            
            # 2. Форматы с "пр." после названия проспекта
            # "пр., д.125, корп.3, кв.30"
            r'пр\.,\s*д\.?\s*(\d+),\s*корп\.?\s*(\d+),\s*кв\.?\s*(\d+)',
            
            # 3. Форматы с тремя компонентами через запятые
            # "д. 84, корп. 3, кв. 124"
            r'\s*д\.?\s*(\d+),\s*корп\.?\s*(\d+),\s*кв\.?\s*(\d+)',
            
            # "д.7, корп. 1, кв. 803" - без пробела после д.
            r'\s*д\.?(\d+),\s*корп\.?\s*(\d+),\s*кв\.?\s*(\d+)',
            
            # "д.6, к.1, кв.34" - с сокращением "к." вместо "корп."
            r'\s*д\.?\s*(\d+),\s*к\.?\s*(\d+),\s*кв\.?\s*(\d+)',
            
            # 4. Форматы с литерой и квартирой
            # "д. 6А кв. 31" - дом с литерой слитно и квартирой
            r'\s*д\.?\s*(\d+)([А-Яа-я])\s+кв\.?\s*(\d+)',
            
            # "д. 30/А кв. 85" - дом с литерой через слэш и квартирой
            r'\s*д\.?\s*(\d+)\/([А-Яа-я])\s+кв\.?\s*(\d+)',
            
            # "д. 24-А кв. 50" - дом-литера через дефис и квартира
            r'\s*д\.?\s*(\d+)-([А-Яа-я])\s+кв\.?\s*(\d+)',
            
            # 5. Форматы с дробным номером дома/корпуса
            # "д. 11/16, кв. 54"
            r'\s*д\.?\s*(\d+)\/(\d+),\s*кв\.?\s*(\d+)',
            
            # 6. Форматы с домом и квартирой через запятую
            # "д. 130, кв. 231"
            r'\s*д\.?\s*(\d+),\s*кв\.?\s*(\d+)',
            
            # "д.14, кв.5" - без пробелов после д. и кв.
            r'\s*д\.?(\d+),\s*кв\.?(\d+)',
            
            # 7. Остальные форматы с тремя компонентами
            # д.19 корп. 3 кв.12
            r'\s*д\.?\s+(\d+)\s+корп\.?\s+(\d+)\s+кв\.?\s+(\d+)',
            
            # д. 58 к. 1 кв. 28, д. 130 к. 1 кв. 390
            r'\s*д\.?\s+(\d+)\s+к\.?\s+(\d+)\s+кв\.?\s+(\d+)',
            
            # д.14 кор.1 кв.204
            r'\s*д\.?\s*(\d+)\s*кор\.?\s*(\d+)\s*кв\.?\s*(\d+)',
            
            # д. 48/3 кв. 87
            r'\s*д\.?\s+(\d+)\/(\d+)\s+кв\.?\s+(\d+)',
            
            # д.19 корп. 3 кв.12 - без пробелов
            r'\s*д\.?(\d+)корп\.?(\d+)кв\.?(\d+)',
            
            # д.14кор.1кв.204 - без пробелов
            r'\s*д\.?(\d+)кор\.?(\d+)кв\.?(\d+)',
            
            # д.58к.1кв.28 - без пробелов
            r'\s*д\.?(\d+)к\.?(\d+)кв\.?(\d+)',
            
            # 8. Форматы с тремя компонентами и специальными символами
            # "106—1-112" - длинное тире в формате дом-корпус-квартира
            r'\s*(\d+)—(\d+)-(\d+)',
            
            # дом-корпус-квартира: 114-4-35
            r'\s*(\d+)-(\d+)-(\d+)',
            
            # "14/3-82", "8/1-207" - дом/корпус-квартира
            r'\s*(\d+)\/(\d+)-(\d+)',
            
            # 9. Форматы с двумя компонентами
            # дом квартира: д. 14 кв. 5
            r'\s*д\.?\s+(\d+)\s+(?:кв\.?|квартира)\s*(\d+)',
            
            # дом квартира: 14 кв. 5 (без д.)
            r'\s*(\d+)\s+(?:кв\.?|квартира)\s*(\d+)',
            
            # дом/корпус квартира: 8/4 кв.34
            r'\s*(\d+)\/(\d+)\s+(?:кв\.?|квартира)\s*(\d+)',
            
            # ", 99-110" - дом-квартира с возможной запятой в начале
            r'(?:,\s*)?(\d+)-(\d+)(?!\d)(?:\s|$|,)',
            
            # 10. Дом с корпусом без квартиры
            # дом корпус: д. 19 корп. 3
            r'\s*д\.?\s+(\d+)\s+(?:корп\.?|кор\.?|к\.?)\s*(\d+)(?!\s*кв)',
            
            # дом корпус: 19 корп. 3 (без д.)
            r'\s*(\d+)\s+(?:корп\.?|кор\.?|к\.?)\s*(\d+)(?!\s*кв)',
            
            # 11. Самые общие форматы
            # просто номер дома с д.: д. 15 
            r'\s*д\.?\s+(\d+)(?!\d)(?!\s*-\d+)(?!\s*\/\d+)(?!\s*(?:корп|кор|к))(?!\s*кв)',
            
            # просто номер дома: 15
            r'\s*(\d+)(?!\d)(?!\s*-\d+)(?!\s*\/\d+)(?!\s*(?:корп|кор|к))(?!\s*кв)'
        ]

    def extract_phone(self, text):
        """
        Извлекает телефонный номер из текста
        """
        if not text or text is None:
            return None, []
            
        text = str(text)
        
        # Находим все позиции цифр в тексте
        digits_with_positions = [(i, c) for i, c in enumerate(text) if c.isdigit()]
        
        if len(digits_with_positions) < 10:
            return None, []  # Недостаточно цифр для номера телефона
        
        # Берем последние 10 цифр с их позициями
        last_10_digits_with_positions = digits_with_positions[-10:]
        
        # Проверяем, что между цифрами нет слишком больших разрывов
        max_gap = 3  # Максимально допустимый промежуток между последовательными цифрами
        
        for i in range(1, len(last_10_digits_with_positions)):
            prev_pos = last_10_digits_with_positions[i-1][0]
            curr_pos = last_10_digits_with_positions[i][0]
            
            if curr_pos - prev_pos > max_gap:
                # Слишком большой разрыв, вероятно, это не телефон
                return None, []
        
        # Извлекаем только цифры из последних 10 позиций
        last_10_digits = ''.join(digit for _, digit in last_10_digits_with_positions)
        
        # Начальная и конечная позиции 10 цифр в тексте
        start_pos = last_10_digits_with_positions[0][0]
        end_pos = last_10_digits_with_positions[-1][0] + 1
        
        # Проверяем, есть ли "8" перед номером
        raw_phone = last_10_digits
        if len(digits_with_positions) > 10:
            # Позиция и значение 11-й цифры с конца
            eleventh_pos, eleventh_digit = digits_with_positions[-11]
            
            # Проверяем, что это "8" и она расположена непосредственно перед началом основных 10 цифр
            # или отделена только допустимыми разделителями
            if eleventh_digit == '8' and (start_pos - eleventh_pos) <= max_gap:
                raw_phone = eleventh_digit + last_10_digits
                start_pos = eleventh_pos  # Включаем "8" в диапазон
        
        # Извлекаем оригинальный текст телефона
        original_phone_text = text[start_pos:end_pos]
        
        return raw_phone, [original_phone_text]

    def extract_address(self, text):
        """
        Извлекает адрес из текста согласно новым жестким правилам.
        Обрабатывает только улицы из словаря и дома/квартиры, указанные после этих улиц.
        """
        if not text or text is None:
            return None, None
            
        text = str(text).strip()
        
        # Предобработка: заменяем длинное тире на обычное
        text = text.replace('—', '-')
        
        # Создаем регулярное выражение для поиска улиц
        # Сортируем названия улиц по длине, чтобы сначала искать более длинные
        sorted_streets = sorted(self.street_names, key=len, reverse=True)
        street_pattern = '|'.join(map(re.escape, sorted_streets))
        
        # Поиск названия улицы
        street_match = re.search(fr'({street_pattern})', text, re.IGNORECASE)
        if not street_match:
            return None, None
            
        street_name = street_match.group(1)
        street_pos = street_match.start()
        
        # Проверка пробелов до и после названия улицы
        # и удаление префиксов (ул., пр., СПб и т.д.)
        
        # Проверяем, есть ли префиксы улиц перед названием
        prefix_text = text[:street_pos]
        prefix_pattern = '|'.join(self.street_prefixes)
        prefix_match = re.search(fr'({prefix_pattern})\s*$', prefix_text, re.IGNORECASE)
        
        # Ищем указание города перед префиксом улицы
        city_pattern = '|'.join(map(re.escape, self.city_prefixes))
        city_text = text[:street_pos]
        city_match = re.search(fr'({city_pattern})[,\s]*$', city_text, re.IGNORECASE)
        
        # Позиция начала адреса (для определения оригинального адреса)
        start_pos = street_pos
        if prefix_match:
            start_pos = prefix_match.start()
            if city_match and city_match.end() == prefix_match.start():
                start_pos = city_match.start()
        elif city_match:
            start_pos = city_match.start()
        
        # Ищем информацию о доме после названия улицы
        house_text = text[street_match.end():]
        
        house_info = ""
        original_house_text = ""
        
        for pattern in self.house_patterns:
            house_match = re.search(pattern, house_text, re.IGNORECASE)
            
            if house_match:
                original_house_text = house_match.group(0)
                house_digits = house_match.groups()
                
                # Форматируем информацию о доме в зависимости от шаблона
                if len(house_digits) == 4:
                    # НОВОЕ ПРАВИЛО для "д. 8/3/А кв. 189" => "8-3А-189"
                    if re.match(r'[А-Яа-я]', house_digits[2]) and pattern.find(r'\/(\d+)\/([А-Яа-я])') != -1:
                        house_info = f"{house_digits[0]}-{house_digits[1]}{house_digits[2].upper()}-{house_digits[3]}"
                    # Проверяем, является ли третья группа литерой
                    elif re.match(r'[А-Яа-я]', house_digits[2]):
                        # Форматы с домом, корпусом, литерой, квартирой
                        # "д. 14, корп. 1, лит. А, кв. 93" -> "14-1А-93"
                        # "27-2-А-17" -> "27-2А-17"
                        house_info = f"{house_digits[0]}-{house_digits[1]}{house_digits[2].upper()}-{house_digits[3]}"
                    else:
                        # Общий случай для четырех компонентов
                        house_info = f"{house_digits[0]}-{house_digits[1]}-{house_digits[2]}-{house_digits[3]}"
                elif len(house_digits) == 3:
                    # НОВОЕ ПРАВИЛО для "54-а-121" => "54А-121"
                    if re.match(r'[а-я]', house_digits[1]) and pattern.find(r'-([а-я])-') != -1:
                        house_info = f"{house_digits[0]}{house_digits[1].upper()}-{house_digits[2]}"
                    # НОВОЕ ПРАВИЛО для "110- А-422" => "110А-422" и для "16 лит. А кв. 43" => "16А-43"
                    elif re.match(r'[А-Яа-я]', house_digits[1]) and (pattern.find(r'-\s*([А-Яа-я])-') != -1 or 
                                                                    pattern.find(r'лит\.?\s+([А-Яа-я])') != -1):
                        house_info = f"{house_digits[0]}{house_digits[1].upper()}-{house_digits[2]}"
                    # Проверяем, является ли вторая группа литерой
                    elif re.match(r'[А-Яа-я]', house_digits[1]):
                        # Форматы с домом, литерой, квартирой
                        # "д. 6А кв. 31" -> "6А-31"
                        # "д. 30/А кв. 85" -> "30А-85"
                        # "д. 24-А кв. 50" -> "24А-50"
                        house_info = f"{house_digits[0]}{house_digits[1].upper()}-{house_digits[2]}"
                    else:
                        # Общий случай для трех компонентов: дом-корпус-квартира
                        house_info = f"{house_digits[0]}-{house_digits[1]}-{house_digits[2]}"
                elif len(house_digits) == 2:
                    # дом-квартира или дом-корпус
                    house_info = f"{house_digits[0]}-{house_digits[1]}"
                elif len(house_digits) == 1:
                    # просто номер дома
                    house_info = house_digits[0]
                
                break
        
        # Определяем конечную позицию адреса
        end_pos = street_match.end()
        if original_house_text:
            house_start = text.find(original_house_text, street_match.end())
            if house_start != -1:
                end_pos = house_start + len(original_house_text)
        
        # Извлекаем полный исходный текст адреса
        original_address = text[start_pos:end_pos].strip()
        
        # Формируем стандартизированный адрес
        # Добавляем правильный тип улицы из словаря
        street_type = self.street_types.get(street_name, "ул.")
        formatted_address = f"{street_type} {street_name}"
        
        if house_info:
            formatted_address += f" {house_info}"
        
        return formatted_address.strip(), original_address

    def clean_other_info(self, text):
        """
        Очищает текст для поля "Иная информация"
        Удаляет специфичные фразы и лишние знаки пунктуации
        """
        if not text or text is None:
            return None
            
        text = str(text).strip()
        
        # Базовая очистка - заменяем множественные пробелы одиночными
        text = re.sub(r'\s+', ' ', text)
        
        # Удаляем специфичные фразы (от более специфичных к более общим)
        text = re.sub(r'Зарг\s+и\s+прож\.?', '', text, flags=re.IGNORECASE)
        text = re.sub(r'Зарег\s+и\s+прож\.?', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\bГражданство\b', '', text, flags=re.IGNORECASE)
        text = re.sub(r'Р\s*Ф\b', '', text)
        text = re.sub(r'\bтел\.?\b', '', text, flags=re.IGNORECASE)
        text = re.sub(r'г\.\s*СПб\b', '', text, flags=re.IGNORECASE)
        text = re.sub(r'г\.\s+', '', text)
        
        # Очищаем от множественных знаков пунктуации
        text = re.sub(r'[,-]+', '', text)
        text = re.sub(r'^\s*[,-]\s*', '', text)
        text = re.sub(r'\s*[,-]\s*$', '', text)
        
        # Финальная очистка пробелов
        text = text.strip()
        
        if not text:
            return None
            
        return text

def process_column_b(excel_file_path):
    """
    Обрабатывает столбец B во всех листах Excel файла
    и распределяет данные по столбцам O, P, Q
    
    Args:
        excel_file_path: путь к Excel файлу после обработки column_b_formatter
    """
    # Создаем экземпляр обработчика адресов
    processor = ImprovedAddressProcessor()
    
    # Открываем файл Excel
    workbook = openpyxl.load_workbook(excel_file_path)
    
    # Считаем общую статистику
    total_stats = {
        'processed_rows': 0,
        'addresses_found': 0,
        'phones_found': 0,
        'other_info_found': 0
    }
    
    # Словарь для хранения статистики по каждому листу
    sheet_stats = {}
    
    # Обрабатываем каждый лист в Excel файле
    for sheet in workbook.worksheets:
        # Инициализируем статистику для текущего листа
        sheet_stats[sheet.title] = {
            'processed_rows': 0,
            'addresses_found': 0,
            'phones_found': 0,
            'other_info_found': 0
        }
        
        # Обрабатываем каждую строку текущего листа
        for row in range(1, sheet.max_row + 1):
            cell_b = sheet.cell(row=row, column=2)  # Столбец B
            
            # Если ячейка не пустая, обрабатываем ее
            if cell_b.value and str(cell_b.value).strip():
                raw_text = str(cell_b.value).strip()
                total_stats['processed_rows'] += 1
                sheet_stats[sheet.title]['processed_rows'] += 1
                
                # 1. Сначала извлекаем телефон
                phone, original_phone_texts = processor.extract_phone(raw_text)
                
                # 2. Удаляем телефон из исходного текста
                text_without_phone = raw_text
                if original_phone_texts:
                    for phone_text in original_phone_texts:
                        text_without_phone = text_without_phone.replace(phone_text, '')
                    total_stats['phones_found'] += 1
                    sheet_stats[sheet.title]['phones_found'] += 1
                
                # 3. Извлекаем адрес из текста без телефона
                formatted_address, original_address = processor.extract_address(text_without_phone)
                
                # 4. Определяем иное - всё, что осталось после удаления телефона и адреса
                other_info = text_without_phone
                if original_address:
                    other_info = other_info.replace(original_address, '')
                    total_stats['addresses_found'] += 1
                    sheet_stats[sheet.title]['addresses_found'] += 1
                
                # Очищаем результат
                other_info = processor.clean_other_info(other_info)
                if other_info:
                    total_stats['other_info_found'] += 1
                    sheet_stats[sheet.title]['other_info_found'] += 1
                
                # Получаем ячейки в столбцах O, P, Q
                cell_o = sheet.cell(row=row, column=15)  # Столбец O для адреса
                cell_p = sheet.cell(row=row, column=16)  # Столбец P для телефона
                cell_q = sheet.cell(row=row, column=17)  # Столбец Q для иного
                
                # Заполняем ячейки
                if formatted_address:
                    cell_o.value = formatted_address
                
                if phone:
                    cell_p.value = phone
                
                if other_info:
                    cell_q.value = other_info
                
                # Очищаем столбец B, если извлекли что-то полезное
                if formatted_address or phone or other_info:
                    cell_b.value = None
    
    # Сохраняем результат
    output_path = excel_file_path
    workbook.save(output_path)
    
    # Выводим статистику по каждому листу
    print(f"Обработка завершена.")
    print("\nСтатистика по листам:")
    for sheet_name, stats in sheet_stats.items():
        print(f"\nЛист '{sheet_name}':")
        print(f"  Обработано строк: {stats['processed_rows']}")
        print(f"  Найдено адресов: {stats['addresses_found']}")
        print(f"  Найдено телефонов: {stats['phones_found']}")
        print(f"  Найдено дополнительной информации: {stats['other_info_found']}")
    
    # Выводим общую статистику
    print("\nОбщая статистика:")
    print(f"  Всего обработано строк: {total_stats['processed_rows']}")
    print(f"  Всего найдено адресов: {total_stats['addresses_found']}")
    print(f"  Всего найдено телефонов: {total_stats['phones_found']}")
    print(f"  Всего найдено дополнительной информации: {total_stats['other_info_found']}")
    
    return output_path

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        output_path = process_column_b(file_path)
        print(f"Файл успешно обработан и сохранен: {output_path}")
    else:
        print("Пожалуйста, укажите путь к файлу как аргумент командной строки.")
        print("Пример: python b_column_parser.py путь_к_файлу.xlsx")